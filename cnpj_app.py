import streamlit as st
import requests
import re
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── Configuração da página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Consulta CNPJ - Receita Federal",
    page_icon="🏢",
    layout="centered",
)

st.markdown("""
<style>
    .main { background-color: #f8f9fa; }
    .stProgress > div > div { background-color: #1f77b4; }
</style>
""", unsafe_allow_html=True)

# ── Cabeçalho ───────────────────────────────────────────────────────────────
st.title("🏢 Consulta CNPJ")
st.caption("Consulta situação cadastral na Receita Federal via BrasilAPI")
st.divider()

# ── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("ℹ️ Como usar")
    st.markdown("""
    **Opção 1 — Excel:**
    1. Arquivo **.xlsx** com CNPJs na coluna A (linha 2 em diante)
    2. Faça upload e clique em **Iniciar Consulta**

    **Opção 2 — Lista:**
    1. Cole os CNPJs na caixa de texto (um por linha)
    2. Clique em **Iniciar Consulta**
    """)
    st.divider()
    st.markdown("**API:** BrasilAPI (gratuita, sem limite fixo)")
    st.markdown("**Formato aceito:** com ou sem pontuação")
    st.caption("Ex: 12.345.678/0001-99 ou 12345678000199")

# ── Funções auxiliares ───────────────────────────────────────────────────────
def formatar_cnpj(cnpj: str) -> str:
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def cor_situacao(situacao: str) -> str:
    s = situacao.upper()
    if s == "ATIVA":
        return "🟢"
    elif s == "BAIXADA":
        return "🔴"
    elif s in ("SUSPENSA", "INAPTA"):
        return "🟡"
    else:
        return "⚪"

def extrair_cnpjs(raws: list) -> tuple[list, list]:
    validos, invalidos = [], []
    for raw in raws:
        limpo = re.sub(r'\D', '', str(raw).strip())
        if len(limpo) == 14:
            validos.append(limpo)
        elif limpo:
            invalidos.append(raw)
    return validos, invalidos

def gerar_excel(dados: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    cabecalhos = ["CNPJ", "Nome Empresarial", "Situação", "Município", "UF", "Atividade Principal"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cores = {
        "ATIVA":    "C6EFCE",
        "BAIXADA":  "FFC7CE",
        "SUSPENSA": "FFEB9C",
        "INAPTA":   "FFEB9C",
    }

    for i, row in enumerate(dados, 2):
        situacao = str(row.get("situacao", "")).upper()
        fill = PatternFill(start_color=cores.get(situacao, "F2F2F2"),
                           end_color=cores.get(situacao, "F2F2F2"),
                           fill_type="solid")
        valores = [
            row.get("cnpj_fmt", ""),
            row.get("nome", ""),
            row.get("situacao", ""),
            row.get("municipio", ""),
            row.get("uf", ""),
            row.get("atividade", ""),
        ]
        for col, val in enumerate(valores, 1):
            ws.cell(row=i, column=col, value=val).fill = fill

    for col, larg in enumerate([22, 45, 12, 20, 6, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def executar_consulta(cnpjs: list):
    resultados = []
    progress_bar = st.progress(0, text="Iniciando consulta...")
    tabela_box   = st.empty()

    for i, cnpj in enumerate(cnpjs):
        cnpj_fmt = formatar_cnpj(cnpj)
        progress_bar.progress(i / len(cnpjs), text=f"Consultando {i+1}/{len(cnpjs)}: {cnpj_fmt}")

        try:
            url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
            resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)

            if resp.status_code == 200:
                data = resp.json()
                atividade = data.get("cnae_fiscal_descricao", "")
                resultados.append({
                    "cnpj_fmt":  cnpj_fmt,
                    "nome":      data.get("razao_social", "N/A"),
                    "situacao":  data.get("descricao_situacao_cadastral", "Desconhecida"),
                    "municipio": data.get("municipio", ""),
                    "uf":        data.get("uf", ""),
                    "atividade": atividade,
                })
            elif resp.status_code == 404:
                resultados.append({"cnpj_fmt": cnpj_fmt, "nome": "Não encontrado",
                                   "situacao": "Não encontrado", "municipio": "", "uf": "", "atividade": ""})
            else:
                resultados.append({"cnpj_fmt": cnpj_fmt, "nome": "Erro",
                                   "situacao": f"HTTP {resp.status_code}", "municipio": "", "uf": "", "atividade": ""})
        except Exception as e:
            resultados.append({"cnpj_fmt": cnpj_fmt, "nome": "Erro",
                               "situacao": str(e)[:40], "municipio": "", "uf": "", "atividade": ""})

        # Tabela parcial
        df_parcial = pd.DataFrame([{
            "CNPJ":      r["cnpj_fmt"],
            "Nome":      r["nome"],
            "Situação":  f"{cor_situacao(r['situacao'])} {r['situacao']}",
            "Município": r.get("municipio", ""),
            "UF":        r.get("uf", ""),
        } for r in resultados])
        tabela_box.dataframe(df_parcial, use_container_width=True, hide_index=True)

    progress_bar.progress(1.0, text="✅ Consulta finalizada!")
    return resultados

def mostrar_resultados(resultados: list):
    st.divider()
    st.subheader("📊 Resumo")

    contagem = {}
    for r in resultados:
        s = r["situacao"].upper()
        contagem[s] = contagem.get(s, 0) + 1

    cols = st.columns(len(contagem))
    for idx, (sit, qtd) in enumerate(sorted(contagem.items())):
        cols[idx].metric(f"{cor_situacao(sit)} {sit.title()}", qtd)

    st.divider()
    excel_bytes = gerar_excel(resultados)
    nome_arquivo = f"resultado_cnpj_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="📥 Baixar resultado em Excel",
        data=excel_bytes,
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

# ── Abas de entrada ──────────────────────────────────────────────────────────
aba_excel, aba_lista = st.tabs(["📂 Importar Excel", "📋 Colar Lista"])

# ── ABA 1: Excel ─────────────────────────────────────────────────────────────
with aba_excel:
    arquivo = st.file_uploader("Selecione o arquivo Excel com os CNPJs", type=["xlsx"])

    if arquivo:
        try:
            df_entrada = pd.read_excel(arquivo, header=0, usecols=[0], dtype=str)
            df_entrada.columns = ["CNPJ"]
            raws = df_entrada["CNPJ"].dropna().tolist()
            cnpjs, invalidos = extrair_cnpjs(raws)

            col1, col2, col3 = st.columns(3)
            col1.metric("📋 Total no arquivo", len(raws))
            col2.metric("✅ Válidos", len(cnpjs))
            col3.metric("❌ Inválidos", len(invalidos))


            if invalidos:
                with st.expander(f"⚠️ {len(invalidos)} CNPJs inválidos"):
                    st.write(invalidos)

            if st.button("🚀 Iniciar Consulta", type="primary", use_container_width=True, key="btn_excel"):
                if cnpjs:
                    resultados = executar_consulta(cnpjs)
                    mostrar_resultados(resultados)
                else:
                    st.warning("Nenhum CNPJ válido encontrado.")

        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")
    else:
        st.info("👆 Faça o upload de um arquivo Excel para começar.")

# ── ABA 2: Lista colada ───────────────────────────────────────────────────────
with aba_lista:
    st.markdown("Cole os CNPJs abaixo, **um por linha**. Aceita com ou sem pontuação.")
    texto = st.text_area(
        "CNPJs",
        height=200,
        placeholder="12.345.678/0001-99\n98765432000100\n11.222.333/0001-81",
        label_visibility="collapsed",
    )

    if texto.strip():
        raws = [linha.strip() for linha in texto.strip().splitlines() if linha.strip()]
        cnpjs, invalidos = extrair_cnpjs(raws)

        col1, col2, col3 = st.columns(3)
        col1.metric("📋 Total inseridos", len(raws))
        col2.metric("✅ Válidos", len(cnpjs))
        col3.metric("❌ Inválidos", len(invalidos))

        minutos, segundos = divmod(len(cnpjs) * 21, 60)
        st.info(f"⏱️ Tempo estimado: ~{minutos}min {segundos}s")

        if invalidos:
            with st.expander(f"⚠️ {len(invalidos)} CNPJs inválidos"):
                st.write(invalidos)

        if st.button("🚀 Iniciar Consulta", type="primary", use_container_width=True, key="btn_lista"):
            if cnpjs:
                resultados = executar_consulta(cnpjs)
                mostrar_resultados(resultados)
            else:
                st.warning("Nenhum CNPJ válido encontrado.")
    else:
        st.info("👆 Cole os CNPJs acima para começar.")
